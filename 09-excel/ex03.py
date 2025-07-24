"""
성적표 읽기 + 평균 점수와 최대 득점자 계산하여 맨 아래 행에 추가
"""

import openpyxl

# 엑셀 파일 열기
workbook = openpyxl.load_workbook('./09-excel/성적표.xlsx')
sheet = workbook.active

print("📊 원본 성적표:")
print("-" * 50)

# 성적 데이터를 저장할 리스트
students = []  # 학생 데이터
subjects = ['국어', '영어', '수학']  # 과목명

# 엑셀에서 데이터 읽기
for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
    filtered_row = []
    for cell in row:
        if cell is not None:
            filtered_row.append(cell)
    
    if filtered_row:
        print(filtered_row)
        
        # 헤더가 아닌 학생 데이터면 저장
        if row_num > 1 and len(filtered_row) >= 5:
            student_data = {
                '번호': filtered_row[0],
                '이름': filtered_row[1],
                '국어': filtered_row[2],
                '영어': filtered_row[3],
                '수학': filtered_row[4]
            }
            students.append(student_data)

print("\n" + "="*50)
print("📈 계산 결과:")
print("="*50)

# 1. 과목별 평균 계산
korean_sum = sum(student['국어'] for student in students)
english_sum = sum(student['영어'] for student in students)
math_sum = sum(student['수학'] for student in students)
student_count = len(students)

korean_avg = korean_sum / student_count
english_avg = english_sum / student_count
math_avg = math_sum / student_count

print(f"📊 과목별 평균:")
print(f"   국어 평균: {korean_avg:.1f}점")
print(f"   영어 평균: {english_avg:.1f}점")
print(f"   수학 평균: {math_avg:.1f}점")

# 2. 최대 득점자 찾기
max_score = 0
top_student = ""

print(f"\n🎯 개별 총점:")
for student in students:
    total = student['국어'] + student['영어'] + student['수학']
    print(f"   {student['이름']}: {total}점")
    
    if total > max_score:
        max_score = total
        top_student = student['이름']

print(f"\n🏆 최대 득점자: {top_student} ({max_score}점)")

# 3. 엑셀에 결과 추가
print(f"\n📝 엑셀에 결과 추가중...")

# 현재 마지막 행 찾기
last_row = sheet.max_row

# 평균 점수 행 추가
avg_row = last_row + 2  # 한 줄 띄우고 추가
sheet.cell(row=avg_row, column=1, value="평균")
sheet.cell(row=avg_row, column=2, value="")
sheet.cell(row=avg_row, column=3, value=round(korean_avg, 1))
sheet.cell(row=avg_row, column=4, value=round(english_avg, 1))
sheet.cell(row=avg_row, column=5, value=round(math_avg, 1))

# 최대 득점자 행 추가
top_row = avg_row + 1
sheet.cell(row=top_row, column=1, value="최고점")
sheet.cell(row=top_row, column=2, value=top_student)
sheet.cell(row=top_row, column=3, value="")
sheet.cell(row=top_row, column=4, value="")
sheet.cell(row=top_row, column=5, value=max_score)

# 새 파일로 저장
new_filename = './09-excel/성적표_결과.xlsx'
workbook.save(new_filename)
workbook.close()

print(f"✅ 결과가 '{new_filename}' 파일에 저장되었습니다!")

print(f"\n📋 추가된 내용:")
print(f"   평균 행: ['평균', '', {korean_avg:.1f}, {english_avg:.1f}, {math_avg:.1f}]")
print(f"   최고점 행: ['최고점', '{top_student}', '', '', {max_score}]")
