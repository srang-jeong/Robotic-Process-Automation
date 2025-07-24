import streamlit as st
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import time
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill, Alignment


# ─────────────────────────────
# 1. 페이지 설정 및 커스텀 CSS
st.set_page_config(
    page_title="🏠 부동산114 전월세 대시보드",
    page_icon="🏠",
    layout="wide"
)
st.markdown("""
<style>
  .reportview-container {background-color: #f5f5f5;}
  .sidebar .sidebar-content {background-color: #ffffff;}
  h1, h2, h3 {color: #2c3e50; font-weight:800;}
  .stMetric-value {font-size:2.5rem !important; color:#e74c3c;}
  .stButton>button {background-color:#3498db; color:white;}
  .stButton>button:hover {background-color:#2980b9;}
</style>
""", unsafe_allow_html=True)

st.title("🏠 부동산114 전월세 매물 대시보드")

# ────────────────
# 2. 크롤링 함수
def 크롤링_부동산114():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)")
    driver = webdriver.Chrome(options=options)
    items = []
    try:
        # 1~5페이지 반복
        for page in range(1, 6):
            url = (
                "https://www.r114.com/?_c=memul&_m=p10"
                "&area_code=41111"
                "&deal_gubun=type3_3"
                f"&page={page}"
            )
            driver.get(url)
            time.sleep(2)
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            # 모든 <li> 매물 루프 (탐색 범위 줄임)
            for li in soup.select('ul.list_article.Best > li'):
                try:
                    # 전월세 구분(월세/전세)
                    kind_tag = li.select_one('strong.tit_a em span.tag_comm2')
                    전월세구분 = kind_tag.get_text(strip=True) if kind_tag else ''

                    # 보증금 및 월세 (e.g., 2,000 / 45 만원)
                    em = li.select_one('strong.tit_a em')
                    보증금, 월세 = None, None
                    if em:
                        # "2,000 / 45 만원" 에서 분리
                        price_parts = em.get_text(" ", strip=True).replace("만원","").split('/')
                        if len(price_parts) == 2:
                            보증금 = int(price_parts[0].replace(",", "").strip())
                            월세_text = price_parts[1].strip()
                            # 월세 부분이 맨 끝에 숫자 뒤에 단위가 붙을 수 있으니 숫자만 추출
                            월세 = int(''.join(filter(str.isdigit, 월세_text)))
                        else:
                            # 혹시 형태가 다르면 보증금만(전세)
                            보증금 = int(''.join(filter(str.isdigit, price_parts[0])))
                            월세 = 0

                    # 주소
                    addr_tag = li.select_one('span.txt p')
                    주소 = addr_tag.get_text(strip=True) if addr_tag else ''
                    items.append({
                        '전월세구분': 전월세구분,
                        '보증금': 보증금,
                        '월세': 월세,
                        '주소': 주소
                    })
                except Exception as e:
                    continue
        driver.quit()
        return pd.DataFrame(items)
    except Exception as e:
        driver.quit()
        st.error(f"크롤링 또는 파싱 중 문제 발생: {e}")
        return pd.DataFrame()
# ────────────────
# 3. 데이터 전처리 및 통계
def 데이터_정제(df):
    if df is None or df.empty:
        return pd.DataFrame()
    df['보증금'] = pd.to_numeric(df['보증금'], errors='coerce').fillna(0)
    df['월세'] = pd.to_numeric(df['월세'], errors='coerce').fillna(0)
    df.drop_duplicates(['주소','보증금','월세'], inplace=True)
    return df

def 통계_분석(df):
    if df is None or df.empty:
        return {}
    return {
        '총매물수':    len(df),
        '전세매물':    len(df[df['전월세구분']=='전세']),
        '월세매물':    len(df[df['전월세구분']=='월세']),
        '평균보증금':  f"{df['보증금'].mean():,.0f}만원",
        '평균월세':    f"{df[df['월세']>0]['월세'].mean():,.0f}만원",
        '최저보증금':  f"{df['보증금'].min():,.0f}만원" if not df.empty else '0만원',
        '최고보증금':  f"{df['보증금'].max():,.0f}만원" if not df.empty else '0만원'
    }

# ────────────────
# 4. 저장 함수
def CSV_저장(df):
    fn = f"r114_전월세_{datetime.now():%Y%m%d_%H%M}.csv"
    df.to_csv(fn, index=False, encoding='utf-8-sig')
    return fn

def 엑셀_리포트(df):
    fn = f"r114_전월세_{datetime.now():%Y%m%d_%H%M}.xlsx"
    wb = Workbook()
    ws = wb.active; ws.title='매물목록'
    cols = list(df.columns)
    ws.append(cols)
    for idx, _ in enumerate(cols, 1):
        cell = ws.cell(1, idx)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4F81BD')
        cell.alignment = Alignment(horizontal='center')
    for row in df.itertuples(index=False):
        ws.append(row)
    if '보증금' in cols:
        chart = BarChart()
        data = Reference(ws, min_col=cols.index('보증금')+1, min_row=2, max_row=1+len(df))
        chart.add_data(data)
        chart.title = "보증금 분포"
        ws.add_chart(chart, 'G2')
    wb.save(fn)
    return fn

# ────────────────
# 5. Streamlit 메인 앱
def main():
    df = None

    src = st.sidebar.radio("데이터 입력 방식", ["🔄 크롤링", "💾 샘플"])
    if src == "🔄 크롤링":
        if st.sidebar.button("🚀 크롤링 실행"):
            df = 크롤링_부동산114()
        else:
            st.sidebar.info("크롤링 버튼을 눌러주세요")
            st.stop()
    else:  # 샘플 모드
        df = pd.DataFrame([{
            '전월세구분':'월세',
            '보증금':3000,
            '월세':35,
            '주소':'수원시 영통구'
        }])

    # 크롤링 결과/샘플 모두 None 또는 비어있을 경우 안내 후 중단
    if df is None or df.empty:
        st.error("처리할 데이터가 없습니다.")
        st.stop()

    df = 데이터_정제(df)
    stats = 통계_분석(df)
    c1, c2, c3 = st.columns(3)
    c1.metric("🏘️ 총 매물수", stats.get('총매물수', 0))
    c2.metric("🏢 전세 매물", stats.get('전세매물', 0))
    c3.metric("🏠 월세 매물", stats.get('월세매물', 0))

    st.markdown("---")
    st.subheader("📋 매물 목록")
    st.dataframe(df.reset_index(drop=True), use_container_width=True)
    st.markdown("---")
    st.subheader("💾 다운로드")
    col1, col2 = st.columns(2)
    if col1.button("📂 CSV 다운로드"):
        fn = CSV_저장(df)
        st.success(f"CSV 파일 생성: {fn}")
    if col2.button("📂 엑셀 다운로드"):
        fn = 엑셀_리포트(df)
        st.success(f"엑셀 파일 생성: {fn}")

if __name__ == "__main__":
    main()