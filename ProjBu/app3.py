import streamlit as st
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import altair as alt

st.set_page_config(page_title="🏠 수원 오피스텔 매물 대시보드", page_icon="🏢", layout="wide")
st.title("🏠 수원 오피스텔 전월세 대시보드")

# 1. HTML 파일 자동 로드
with open("paste.txt", encoding="utf-8") as f:
    html = f.read()
soup = BeautifulSoup(html, "html.parser")
items = []

for li in soup.select("li"):
    a_tag = li.select_one("a.cont")
    if not a_tag:
        continue
    apt_name = a_tag.select_one('strong.tit_a > span')
    apt_name = apt_name.get_text(strip=True) if apt_name else ''
    # 전월세 구분
    deal_type_tag = a_tag.select_one('strong.tit_a em span.tag_comm2')
    deal_type = deal_type_tag.get_text(strip=True) if deal_type_tag else ''
    # 보증금/월세 (em에서 '월세' 부분 제거 후 금액 추출)
    price_em = a_tag.select_one('strong.tit_a em')
    deposit, rent = 0, 0
    if price_em:
        text_for_money = price_em.get_text(separator="", strip=True)
        # "월세2,000 / 45 만원"
        if deal_type:
            # remove "월세" (or "전세") prefix
            text_for_money = text_for_money.replace(deal_type, "", 1)
        text_for_money = text_for_money.replace("만원", "").replace(" ", "")
        splitted = text_for_money.split("/")
        if len(splitted) == 2:
            try:
                deposit = int(splitted[0].replace(",",""))
                rent = int(splitted[1].replace(",",""))
            except Exception:
                deposit, rent = 0, 0
        elif splitted and splitted[0]:
            try:
                deposit = int(splitted[0].replace(",",""))
                rent = 0
            except Exception:
                deposit, rent = 0, 0
    tag_type = a_tag.select_one('span.tag_comm3')
    listing_type, listing_date = '', ''
    if tag_type:
        em_inside = tag_type.select_one('em')
        if em_inside:
            listing_type = tag_type.get_text(strip=True).replace(em_inside.get_text(strip=True), "")
            listing_date = em_inside.get_text(strip=True)
        else:
            listing_type = tag_type.get_text(strip=True)
    officetel_type = a_tag.select_one('span.txt > strong')
    officetel_type = officetel_type.get_text(strip=True) if officetel_type else ''
    room_area_floor = a_tag.select_one("span.info_memul")
    room_area_floor = room_area_floor.get_text(strip=True) if room_area_floor else ''
    desc = a_tag.select_one("span.txt > em")
    desc = desc.get_text(strip=True) if desc else ''
    addr_tag = a_tag.select_one("span.txt > p")
    addr = addr_tag.get_text(strip=True).replace("위치", "") if addr_tag else ''
    img_tag = a_tag.select_one("div.thumb img")
    img_url = img_tag["src"] if img_tag and img_tag.get("src") else ''
    agent_name, agent_tel = "", ""
    em_tel = li.select_one("em.tel > div")
    if em_tel:
        name_tag = em_tel.select_one("p.name")
        agent_name = name_tag.get_text(strip=True) if name_tag else ''
        tel_tag = em_tel.select_one("span")
        agent_tel = tel_tag.get_text(strip=True).replace("전화번호","") if tel_tag else ''
    items.append({
        "단지명": apt_name,
        "전월세구분": deal_type,
        "보증금": deposit,
        "월세": rent,
        "매물유형": listing_type,
        "매물일자": listing_date,
        "오피스텔유형": officetel_type,
        "방/면적/층수": room_area_floor,
        "상세설명": desc,
        "주소": addr,
        "이미지URL": img_url,
        "중개사명": agent_name,
        "연락처": agent_tel
    })

df = pd.DataFrame(items)
if df.empty:
    st.error("❌ 처리할 매물 데이터가 없습니다.")
    st.stop()

# 2. 데이터 전처리/통계
df['보증금'] = pd.to_numeric(df['보증금'], errors="coerce").fillna(0)
df['월세'] = pd.to_numeric(df['월세'], errors="coerce").fillna(0)
df.drop_duplicates(subset=["단지명","보증금","월세","주소"], inplace=True)
stats = {
    "총매물수": len(df),
    "전세매물": len(df[df["전월세구분"].str.contains("전세")]),
    "월세매물": len(df[df["전월세구분"].str.contains("월세")]),
    "평균보증금": f"{df['보증금'].mean():,.0f}만원",
    "평균월세": f"{df[df['월세']>0]['월세'].mean():,.0f}만원" if (df['월세']>0).any() else "0만원"
}

k1, k2, k3 = st.columns(3)
k1.metric("🏘️ 총 매물", stats["총매물수"])
k2.metric("🏢 전세", stats["전세매물"])
k3.metric("🏠 월세", stats["월세매물"])
st.markdown("---")

st.subheader("📈 가격 분포")
st.altair_chart(
    alt.Chart(df).transform_fold(
      ["보증금","월세"], as_=["항목","금액"]).mark_boxplot(size=35).encode(
        x=alt.X("항목:N", title="항목"),
        y=alt.Y("금액:Q", title="금액(만원)"),
        color="항목:N"
    ).properties(width=380, height=300),
    use_container_width=True
)

st.subheader("🗺️ 동별 매물수 TOP8")
df['동'] = df['주소'].str.extract(r'수원시\s+([^\s]+)동')
dong_count = df['동'].value_counts().head(8).reset_index()
dong_count.columns = ['동','매물수']
st.bar_chart(dong_count.set_index('동'))

st.subheader("✨ 대표 매물 카드")
for idx, row in df.head(5).iterrows():
    col1, col2 = st.columns([1,4])
    with col1:
        if row["이미지URL"]:
            st.image(row["이미지URL"], width=120)
    with col2:
        st.markdown(f"""
**[{row['단지명']}]** {row['전월세구분']}
💰 {int(row['보증금']):,}만원 / 💸 {int(row['월세']):,}만원
📝 {row['상세설명']}
{row['매물유형']} | {row['매물일자']} | {row['방/면적/층수']} | {row['오피스텔유형']}
🏠 {row['주소']}
👤 {row['중개사명'] or ''} | ☎ {row['연락처'] or ''}
""", unsafe_allow_html=True)
    st.markdown("---")

st.subheader("📋 전체 매물 테이블")
show_cols = [
    "단지명","전월세구분","보증금","월세","매물유형","매물일자",
    "오피스텔유형","방/면적/층수","상세설명","주소","중개사명","연락처"
]
st.dataframe(df[show_cols].reset_index(drop=True), use_container_width=True)

def CSV_저장(df):
    fn = f"suwon_r114_{datetime.now():%Y%m%d_%H%M}.csv"
    df.to_csv(fn, index=False, encoding="utf-8-sig")
    return fn

def 엑셀_리포트(df):
    fn = f"suwon_r114_{datetime.now():%Y%m%d_%H%M}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "매물"
    cols = list(df.columns)
    ws.append(cols)
    for idx, _ in enumerate(cols, 1):
        cell = ws.cell(1, idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1565c0")
        cell.alignment = Alignment(horizontal="center")
    for row in df.itertuples(index=False):
        ws.append(row)
    wb.save(fn)
    return fn

st.subheader("💾 다운로드")
c1, c2 = st.columns(2)
if c1.button("📂 CSV 저장"):
    fn = CSV_저장(df)
    st.success(f"CSV 저장됨: {fn}")
if c2.button("📂 엑셀 저장"):
    fn = 엑셀_리포트(df)
    st.success(f"엑셀 저장됨: {fn}")