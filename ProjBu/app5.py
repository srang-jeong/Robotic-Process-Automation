import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

st.set_page_config(page_title="🏠 수원 오피스텔 전월세 대시보드", page_icon="🏢", layout="wide")
st.title("🏠 수원 오피스텔 전월세 매물 대시보드")

CSV_PATH = "suwon_r114_20250722_1607.csv"
try:
    df = pd.read_csv(CSV_PATH, encoding="utf-8")
except FileNotFoundError:
    st.error(f"❌ '{CSV_PATH}' 파일이 현재 폴더에 없습니다. 파일을 먼저 넣어주세요.")
    st.stop()

# 데이터 전처리
df['보증금'] = pd.to_numeric(df['보증금'], errors="coerce").fillna(0)
df['월세'] = pd.to_numeric(df['월세'], errors="coerce").fillna(0)
df.drop_duplicates(subset=["단지명","보증금","월세","주소"], inplace=True)

stats = {
    "총매물수": len(df),
    "전세매물": len(df[df["전월세구분"].str.contains("전세")]),
    "월세매물": len(df[df["전월세구분"].str.contains("월세")]),
}

k1, k2, k3 = st.columns(3)
k1.metric("🏘️ 총 매물", stats["총매물수"])
k2.metric("🏢 전세", stats["전세매물"])
k3.metric("🏠 월세", stats["월세매물"])
st.markdown("---")

st.subheader("✨ 대표 매물 카드")
for idx, row in df.head(6).iterrows():
    col1, col2 = st.columns([1,4])
    with col1:
        if pd.notnull(row.get("이미지URL","")) and str(row["이미지URL"]).strip():
            st.image(row["이미지URL"], width=120)
    with col2:
        # 사이트 링크(상세링크 컬럼 있으면 링크 생성, 없으면 텍스트만)
        r = row
        apt_with_link = r["단지명"]
        if "상세링크" in row and pd.notnull(row["상세링크"]):
            apt_with_link = f'<a href="{row["상세링크"]}" target="_blank" style="color:#1565c0;text-decoration:underline;font-weight:bold">{row["단지명"]}</a>'
        st.markdown(f"""
**[{apt_with_link}]** {r['전월세구분']}  
💰 <b style='color:#1565c0'>{int(r['보증금']):,}만원</b> /
💸 <b style='color:#d84315'>{int(r['월세']):,}만원</b>  
📝 {r['상세설명']}  
{r['매물유형']} | {r['매물일자']} | {r['방/면적/층수']} | {r['오피스텔유형']}  
🏠 {r['주소']}  
👤 {r['중개사명'] or ''} | ☎ {r['연락처'] or ''}
""", unsafe_allow_html=True)
    st.markdown("---")

st.subheader("📋 전체 매물 테이블")
# 단지명에 링크 포함(상세링크가 있으면 하이퍼링크)
if "상세링크" in df.columns:
    df['단지명_링크'] = df.apply(
        lambda r: f'<a href="{r["상세링크"]}" target="_blank">{r["단지명"]}</a>' if pd.notnull(r["상세링크"]) else r["단지명"], axis=1
    )
    show_cols = ["단지명_링크"] + [c for c in df.columns if c not in ["단지명_링크","상세링크","이미지URL"]]
else:
    show_cols = ["단지명"] + [c for c in df.columns if c not in ["단지명","이미지URL"]]
st.markdown(df[show_cols].to_html(escape=False, index=False), unsafe_allow_html=True)

def CSV_저장(df):
    fn = f"suwon_r114_{datetime.now():%Y%m%d_%H%M}.csv"
    save = df.drop(columns=["단지명_링크"], errors="ignore")
    save.to_csv(fn, index=False, encoding="utf-8-sig")
    return fn

def 엑셀_리포트(df):
    fn = f"suwon_r114_{datetime.now():%Y%m%d_%H%M}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "매물"
    xcols = [c for c in df.columns if c not in ['단지명_링크']]
    ws.append(xcols)
    for idx, _ in enumerate(xcols, 1):
        cell = ws.cell(1, idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1565c0")
        cell.alignment = Alignment(horizontal="center")
    for row in df[xcols].itertuples(index=False):
        ws.append(row)
    wb.save(fn)
    return fn

st.subheader("💾 다운로드")
c1, c2 = st.columns(2)
if c1.button("📂 CSV 저장"):
    fn = CSV_저장(df)
    st.success(f"CSV 저장됨: {fn}")
if c2.button("📂 엑셀 저장"):
    fn = 엑셀_리포트(df)
    st.success(f"엑셀 저장됨: {fn}")